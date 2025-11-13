
#######################################################################################
#                 LIPID MAPS REST API + PubChem REST API                              # 
#######################################################################################
# El siguiente código:
    # 1. Lee el Excel con los ID's de LIPID MAPS y PubChem.
    # 2. Filtra los IDs: de LIPID MAPS para metabolitos o proteínas, y los de PubChem. 
    # 3. 
    #  a) Usa la API REST de LIPID MAPS® (/rest/compound/lm_id/{lm_id}/all)  
    #       (https://www.lipidmaps.org/resources/rest) para obtener la InChI e InChIKey de
    #       cada metabolito y las cross-refs de los IDs con KEGG, PubChem, HMDB y ChEBI
    #       (metabolitos) y RefSeq_ID y UniProt en el caso de proteínas.
    #  b) Usa las API REST de PubChem (https://pubchem.ncbi.nlm.nih.gov/docs/rdf-rest) 
    #       para obtener, tanto los datos de la InChI e InChIKey de cada metabolito, 
    #       como las cross-refs de los IDs con KEGG, HMDB y ChEBI
    # 4. Genera una copia del Excel actualizado "_updated.xlsx".
#######################################################################################


# import pandas as pd
# import requests
# import time
# from openpyxl import load_workbook
# import re

# # Columnas de interés
# TARGET_COLS = ["KEGG", "PubChem", "HMDB", "ChEBI", "RefSeq_Id", "UniProt", "InChIKey", "InChI"]

# # Definir UniProt regexp
# uniprot_pattern = r"([OPQ][0-9][A-Z0-9]{3}[0-9]|[A-NR-Z][0-9]([A-Z][A-Z0-9]{2}[0-9]){1,2})"
#     # https://www.uniprot.org/help/accession_numbers


# def fetch_lipidmaps_info(query_id):
#     """Fetch cross-references and InChI info from LipidMaps REST API."""
#     if not query_id:
#         return None

#     # Filtrar tipo de IDs
#     if query_id.startswith("LMP"):
#         ext = "protein/lmp_id" # extensión para queries de LIPID MAPS protein ID
#     elif re.match(uniprot_pattern, query_id):
#         ext = "protein/uniprot_id" # extensión para queries de UniProt protein ID
#     else:
#         ext = "compound/lm_id" # extensión para queries de LIPID MAPS metabolite ID

#     url = f"https://www.lipidmaps.org/rest/{ext}/{query_id}/all"

#     # Consultar la base de datos
#     try:
#         r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
#         r.raise_for_status()
#         data = r.json()
#         if isinstance(data, list) and not data:
#             return None

#         # Caso PROTEIN: múltiples filas, usar solo Row1
#         if isinstance(data, dict) and any(k.startswith('Row') for k in data):
#             first_row = data.get('Row1') or {}
#             return {
#                 "LM_ID": first_row.get("lm_id") or first_row.get("lmp_id") or query_id,
#                 "RefSeq_Id": first_row.get("refseq_id"),
#                 "UniProt": first_row.get("uniprot_id")
#             }    
#         # Caso METABOLITE/PROTEIN: dict plano
#         elif isinstance(data, dict):
#             # extraer LM id
#             lm_from_data = data.get("lm_id") or data.get("lmp_id") or None
#             return {
#                 "LM_ID": lm_from_data if lm_from_data 
#                             else (query_id if query_id.startswith("LM") else None),
#                 "KEGG": data.get("kegg_id"),
#                 "PubChem": data.get("pubchem_cid"),
#                 "HMDB": data.get("hmdb_id"),
#                 "ChEBI": data.get("chebi_id"),
#                 "RefSeq_Id": data.get("refseq_id"),
#                 "UniProt": data.get("uniprot_id"),
#                 "InChIKey": data.get("inchi_key"),
#                 "InChI": data.get("inchi")
#             }
#         else:
#             print(f"[ERROR] Unexpected data format for {query_id}: {type(data)}")
#             return None

#     except (requests.exceptions.RequestException, ValueError) as e: 
#         print(f"[ERROR] No data for {query_id} ({e})") 
#         return None



# def fetch_pubchem_info(pubchem_id):
#     """Fetch compound info from PubChem REST API (using CID)."""
#     try:
#         cid = re.sub(r"[^0-9]", "", pubchem_id)  # mantener solo dígitos
#         if not cid:
#             return None
        
#         # Extraer InChI e InChIKey
#         url_inchi = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/property/InChI,InChIKey/JSON"
#         r_inchi = requests.get(url_inchi, timeout=10)
#         r_inchi.raise_for_status()
#         data_inchi = (
#             r_inchi.json()
#             .get("PropertyTable", {})
#             .get("Properties", [{}])[0]
#         )
#         # Extraer crossreferences (xrefs)
#         url_xrefs = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/xrefs/RegistryID/JSON"
#         r_xrefs = requests.get(url_xrefs, timeout=10)
#         r_xrefs.raise_for_status()
#         data_xrefs = (
#             r_xrefs.json()
#             .get("InformationList", {})
#             .get("Information", [{}])[0]
#             .get("RegistryID", [])
#         )

#         kegg_id = next((x for x in data_xrefs if re.fullmatch(r"C\d{5}", x)), None) # C+5dígitos
#         chebi_id = next((x for x in data_xrefs if x.startswith("CHEBI:")), None)
#         hmdb_id = next((x for x in data_xrefs if x.startswith("HMDB")), None)

#         res = {
#             "PubChem": cid,
#             "KEGG": kegg_id,
#             "HMDB": hmdb_id,
#             "ChEBI": chebi_id,
#             "InChI": data_inchi.get("InChI"),
#             "InChIKey": data_inchi.get("InChIKey")
#         }
#         return res
#     except Exception as e:
#         print(f"[WARN] PubChem no data for {pubchem_id}: {e}")
#         return None



# def fetch_refseq_from_uniprot(uniprot_id):
#     """Fetch RefSeq cross-references from UniProt REST API."""
#     try:
#         url = f"https://rest.uniprot.org/uniprotkb/{uniprot_id}?format=json"
#         r = requests.get(url, timeout=10)
#         r.raise_for_status()
#         data = r.json()

#         refseq_ids = []
#         for xref in data.get("uniProtKBCrossReferences", []):
#             if xref.get("database") == "RefSeq":
#                 refseq_ids.append(xref.get("id"))

#         # Opcional: incluir los IDs de nucleótidos asociados
#         refseq_pairs = []
#         for xref in data.get("uniProtKBCrossReferences", []):
#             if xref.get("database") == "RefSeq":
#                 ref_id = xref.get("id")
#                 nuc = None
#                 for prop in xref.get("properties", []):
#                     if prop.get("key") == "NucleotideSequenceId":
#                         nuc = prop.get("value")
#                 refseq_pairs.append((ref_id, nuc))

#         return {
#             "UniProt": uniprot_id,
#             "RefSeq_IDs": refseq_ids or None,
#             "RefSeq_Pairs": refseq_pairs or None
#         }

#     except Exception as e:
#         print(f"[WARN] UniProt no data for {uniprot_id}: {e}")
#         return None


# #################################### MAIN ######################################
# # 1. Leer Excel
# input_file = "c:/Users/dborrotoa/Desktop/TFM/PathwayBA_list.xlsx"
# # input_file = "C:/Users/deyan/Desktop/BIOINFORMATICA/1TFM/PathwayBA_list.xlsx"
# wb = load_workbook(input_file)
# total_time = []

# for sheet_name in wb.sheetnames:
#     inicio = time.perf_counter()
#     ws = wb[sheet_name]
#     print(f"\n--- Processing sheet: {sheet_name} ---")

#     # Cargar datos de la hoja
#     df = pd.DataFrame(ws.values)
#     df.columns = df.iloc[0]     # Primera fila = header
#     df = df[2:]                 # Saltamos header y nombre del pathway

#     # Filtrado y conteo de IDs 
#     seen_queries = set()
#     lm_ids, other_ids = [], []
#     for id_raw in df["ID"]:
#         id_str = str(id_raw).strip() if pd.notna(id_raw) else '' # si el ID dado no es nulo
#         if not id_str:
#             continue
#         for sub in [s.strip() for s in id_str.split(";") if s.strip()]:
#             if sub in seen_queries:
#                 continue
#             seen_queries.add(sub)
#             if sub.startswith("LM"):
#                 lm_ids.append(sub)
#             else:
#                 other_ids.append(sub)


#     print(f"Found {len(lm_ids)} LIPID MAPS IDs in {sheet_name}.")
#     print(f"Found {len(other_ids)} IDs wich are not from LIPID MAPS in {sheet_name}:{other_ids}")

#     # Consultar LipidMaps  y crear index de búsqueda
#     results = {}
#     for query_id in (lm_ids + other_ids):
#         info = fetch_lipidmaps_info(query_id)
#         if not info:
#             time.sleep(0.25)
#             continue

#         # LM ID
#         lm_key = info.get("LM_ID")
#         if lm_key:
#             results[lm_key] = info

#         # UniProt ID
#         uni_field = info.get("UniProt")
#         if uni_field:
#             for uni in [u.strip() for u in str(uni_field).split(";") if u.strip()]:
#                 results[uni] = info
        
#         # Indexar por la consulta original (por si no hay LM o UniProt)
#         results[query_id] = info
#         time.sleep(0.25)

#     # Cálculo de IDs no crossreferenciados
#     not_crossreferenced = set()
#     for id_raw in df["ID"]:
#         id_str = str(id_raw).strip() if pd.notna(id_raw) else ''
#         if not id_str:
#             continue
#         for sub in [s.strip() for s in id_str.split(";") if s.strip()]:
#             if sub not in results:
#                 not_crossreferenced.add(sub)
    
#     # if not_crossreferenced:
#     #     print(f"--- [{len(not_crossreferenced)}] IDs finished with no crossreference in {sheet_name}: {sorted(not_crossreferenced)}")
#     # else:
#     #     print(f"Todas las IDs cruzadas en {sheet_name}.")


#     # Recuperar información de PubChem y UniProt
#     # Uniprot: [a implementar]
#     external_results = {}
#     uniprot_id = ["Q06520","Q14032", "Q9Y2P5"]
#     for id in uniprot_id:
#         info = fetch_refseq_from_uniprot(id)
#     # PubChem
#     for unknown_id in sorted(not_crossreferenced):
#         if re.match(r"^\d+$", unknown_id):  ## NO ME GUSTA QUE TENEMOS 2 VECES BUSCAR CID CON REGEXP DISTINTAS
#             info = fetch_pubchem_info(unknown_id)
#         else:
#             continue

#         if info:
#             external_results[unknown_id] = info
#         time.sleep(0.3)

#     # Unir resultados externos al diccionario principal
#     results.update(external_results)


#     # Excluir IDs UniProt del resumen
#     not_crossreferenced_final = [
#         x for x in sorted(not_crossreferenced)
#         if x not in results and not re.match(uniprot_pattern, x)
#     ]

#     if not_crossreferenced_final:
#         print(f"--- [{len(not_crossreferenced_final)}] IDs finished with no crossreference in {sheet_name}: {not_crossreferenced_final}")
#     else:
#         print(f"Todas las IDs cruzadas en {sheet_name}.")


#     # Mapear columnas a índices en la hoja Excel
#     header = [cell.value for cell in ws[1]]
#     header_lower = [str(h).lower() if h else "" for h in header]
#     target_col_idx = {}
#     for col in TARGET_COLS:
#         if col.lower() in header_lower:
#             target_col_idx[col] = header_lower.index(col.lower()) + 1

#     if "id" not in header_lower:
#         print(f"[WARN] No 'ID' header in sheet {sheet_name}, skip.")
#         continue
#     id_idx = header_lower.index("id") + 1


#     # Actualizar celdas
#     for row in ws.iter_rows(min_row=2, values_only=False):
#         id_cell = row[id_idx - 1]
#         id_val = str(id_cell.value).strip() if pd.notna(id_cell.value) else ''
#         if not id_val:
#             continue

#         sub_ids = [x.strip() for x in id_val.split(";") if x.strip()]
#         merged_info = {k: [] for k in TARGET_COLS}
#         new_id_val = []

#         for sub_id in sub_ids:
#             info = results.get(sub_id)
#             if not info and re.match(uniprot_pattern, sub_id):
#                 info = results.get(sub_id)
#             if info:
#                 # reemplazar UniProt por LM si existe LM_ID real
#                 lm_real = info.get("LM_ID")
#                 if lm_real:
#                     new_id_val.append(lm_real)
#                 else:
#                     new_id_val.append(sub_id)

#                 # para cada campo target, añadir valores evitando duplicados
#                 for col in TARGET_COLS:
#                     v = info.get(col)
#                     if v:
#                         # si la API devuelve varios dentro del mismo campo, separar por ';'
#                         for part in [p.strip() for p in str(v).split(";") if p.strip()]:
#                             if part not in merged_info[col]:
#                                 merged_info[col].append(part)
#             else:
#                 # si no hay info, conservar el sub-id sin cambios
#                 new_id_val.append(sub_id)

#         # escribir nuevo ID concatenado (mantener orden encontrado)
#         id_cell.value = ";".join(new_id_val)

#         # escribir columnas objetivo concatenadas
#         for col_name, col_idx in target_col_idx.items():
#             vals = merged_info.get(col_name) or []
#             if vals:
#                 ws.cell(row=id_cell.row, column=col_idx, value=";".join(vals))

#     fin = time.perf_counter()
#     total_time.append(fin - inicio)
#     print(f"Tiempo de ejecución: {fin - inicio:.2f} s")

# output_file = input_file.replace(".xlsx", "_updatedB.xlsx")
# wb.save(output_file)
# print(f"Archivo guardado como: {output_file}")
# print(f"Tiempo total: {sum(total_time):.2f} s")


# ########################################### MAIN ############################################
# # Estuctura:
# # main.py
# # ├── read_input_excel()        # Carga datos desde Excel y devuelve {sheet_name: DataFrame}
# # ├── classify_ids()            # Separa LM, UniProt, y otros IDs
# # ├── fetch_lipidmaps_info()    # Busca información y cross-referencias de LIPID MAPS
# # ├── fetch_pubchem_info()      # Busca información y cross-referencias de PubChem
# # ├── fetch_refseq_from_uniprot() # Busca información y cross-referencias de UniProt
# # ├── integrate_crossrefs()     # Une resultados de LipidMaps, PubChem y UniProt
# # ├── update_excel()            # Escribe los resultados en el Excel
# # └── main()                    # Coordina todo

# import pandas as pd
# import requests
# import time
# import re
# from openpyxl import load_workbook


# ########################################### CONFIG ###########################################

# TARGET_COLS = [
#     "KEGG", "PubChem", "HMDB", "ChEBI",
#     "RefSeq_Id", "UniProt", "InChIKey", "InChI"
# ]
# uniprot_pattern = r"([OPQ][0-9][A-Z0-9]{3}[0-9]|[A-NR-Z][0-9]{5})"


# ######################################## LIPID MAPS ##########################################

# def fetch_lipidmaps_info(qid):
#     """Fetch xrefs from LipidMaps by LM_ID."""
#     url = f"https://www.lipidmaps.org/rest/compound/lm_id/{qid}/all"
#     try:
#         r = requests.get(url, timeout=15)
#         if r.status_code != 200:
#             return None
#         data = r.json()
#         # Extract only known reference fields
#         info = {k: data.get(k, None) for k in TARGET_COLS if k in data}
#         info["Name"] = data.get("name", None)
#         info["LM_ID"] = data.get("lm_id", qid)
#         return info
#     except Exception:
#         return None
    

# ######################################## PUBCHEM ############################################

# def fetch_pubchem_info(qid):
#     """Fetch RegistryID ↔ SourceName from PubChem by numeric CID or other ID."""
#     url = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{qid}/xrefs/RegistryID,SourceName/JSON"
#     try:
#         r = requests.get(url, timeout=15)
#         if r.status_code != 200:
#             return None
#         data = r.json()
#         info_list = data.get("InformationList", {}).get("Information", [])
#         if not info_list:
#             return None
#         info = info_list[0]
#         reg = info.get("RegistryID", [])
#         src = info.get("SourceName", [])
#         mapped = dict(zip(reg, src)) if reg and src else {}
#         return {"PubChem": qid, "Registry_Source_Map": mapped}
#     except Exception:
#         return None
    

# #################################### UNIPROT → REFSEQ #######################################

# def fetch_refseq_from_uniprot(uid):
#     """Get RefSeq accessions from UniProt JSON."""
#     url = f"https://rest.uniprot.org/uniprotkb/{uid}.json"
#     try:
#         r = requests.get(url, timeout=15)
#         if r.status_code != 200:
#             return None
#         data = r.json()
#         xrefs = data.get("uniProtKBCrossReferences", [])
#         refseq_ids = [
#             x["id"] for x in xrefs if x.get("database") == "RefSeq"
#         ]
#         return {"UniProt": uid, "RefSeq_IDs": refseq_ids}
#     except Exception:
#         return None

    

# ############################### CROSS-REFERENCE INTEGRATION #################################

# def integrate_crossrefs(df):
#     """Integrate LipidMaps, PubChem, and UniProt cross-references for all IDs."""
#     seen = set()
#     results = {}
#     not_cross = set()

#     lm_ids, other_ids, uniprot_ids = [], [], []

#     # --- Classify IDs
#     for id_raw in df["ID"]:
#         id_str = str(id_raw).strip() if pd.notna(id_raw) else ""
#         if not id_str:
#             continue
#         for sub in [s.strip() for s in id_str.split(";") if s.strip()]:
#             if sub in seen:
#                 continue
#             seen.add(sub)
#             if sub.startswith("LM"):
#                 lm_ids.append(sub)
#             elif re.match(uniprot_pattern, sub):
#                 uniprot_ids.append(sub)
#             else:
#                 other_ids.append(sub)

#     print(f"  LIPID MAPS: {len(lm_ids)}  |  UniProt: {len(uniprot_ids)}  |  Otros: {len(other_ids)}")

#     # --- LipidMaps
#     for qid in lm_ids + other_ids:
#         info = fetch_lipidmaps_info(qid)
#         if not info:
#             not_cross.add(qid)
#             continue
#         results[qid] = info
#         time.sleep(0.2)

#     # --- UniProt → RefSeq
#     for uid in uniprot_ids:
#         info = fetch_refseq_from_uniprot(uid)
#         if info:
#             results[uid] = {
#                 "UniProt": uid,
#                 "RefSeq_Id": ";".join(info["RefSeq_IDs"]) if info["RefSeq_IDs"] else None
#             }
#         else:
#             not_cross.add(uid)
#         time.sleep(0.2)

#     # --- PubChem fallback for numeric IDs
#     for qid in sorted(not_cross.copy()):
#         if re.match(r"^\d+$", qid):
#             info = fetch_pubchem_info(qid)
#             if info:
#                 results[qid] = info
#                 not_cross.discard(qid)
#         time.sleep(0.3)

#     return results, not_cross


# ##################################### EXCEL WRITER ###########################################
# def update_excel(ws, df, results, not_cross):
#     """Write cross-reference results back into Excel sheet."""
#     id_col = list(df.columns).index("ID") + 1
#     existing_cols = [c for c in ws[1] if c.value in TARGET_COLS]
#     col_offset = len(existing_cols)

#     # Find column numbers for TARGET_COLS
#     header = [cell.value for cell in ws[1]]
#     col_map = {col: header.index(col) + 1 if col in header else None for col in TARGET_COLS}

#     for i, id_raw in enumerate(df["ID"], start=3):  # data starts at row 3
#         if pd.isna(id_raw):
#             continue
#         ids = [x.strip() for x in str(id_raw).split(";") if x.strip()]
#         for qid in ids:
#             info = results.get(qid)
#             if not info:
#                 continue
#             for key, value in info.items():
#                 if key not in TARGET_COLS:
#                     continue
#                 col = col_map.get(key)
#                 if not col:
#                     continue
#                 ws.cell(row=i, column=col, value=value)

#     # Mark non-crossreferenced IDs
#     not_cross_str = ";".join(sorted(not_cross)) if not_cross else ""
#     ws.cell(row=2, column=col_offset + 2, value="Not_CrossReferenced")
#     ws.cell(row=3, column=col_offset + 2, value=not_cross_str)


# ##################################### MAIN EXECUTION #########################################

# def main(input_file):
#     wb = load_workbook(input_file)
#     total_time = []

#     for sheet_name in wb.sheetnames:
#         print(f"\n--- Processing sheet: {sheet_name} ---")
#         start = time.perf_counter()

#         ws = wb[sheet_name]
#         df = pd.DataFrame(ws.values)
#         df.columns = df.iloc[0]
#         df = df[2:]  # skip first two rows (headers/meta)

#         results, not_cross = integrate_crossrefs(df)
#         update_excel(ws, df, results, not_cross)

#         end = time.perf_counter()
#         total_time.append(end - start)
#         print(f"  Tiempo de ejecución: {end - start:.2f} s")

#     output_file = input_file.replace(".xlsx", "_updated.xlsx")
#     wb.save(output_file)
#     print(f"\nArchivo guardado como: {output_file}")
#     print(f"Tiempo total: {sum(total_time):.2f} s")


# ##################################### ENTRY POINT ##########################################

# if __name__ == "__main__":
#     input_excel = "c:/Users/dborrotoa/Desktop/TFM/PathwayBA_list.xlsx"  
#     main(input_excel)


# # input_file = "c:/Users/dborrotoa/Desktop/TFM/PathwayBA_list.xlsx"
# # input_file = "C:/Users/deyan/Desktop/BIOINFORMATICA/1TFM/PathwayBA_list.xlsx"


#############################################################################################

#############################################################################################

import pandas as pd
import requests
import time
import logging
import re
from openpyxl import load_workbook

# Configurar logging para mejor depuración (registra mensajes importantes y errores)
logging.basicConfig(level=logging.INFO, format="%(levelname)s:%(message)s")

# Columnas de interés
TARGET_COLS = ["KEGG", "PubChem", "HMDB", "ChEBI", "RefSeq_Id", "UniProt", "InChIKey", "InChI"]

# Patron de ID UniProt (verificado según reglas oficiales)
UNIPROT_PATTERN = r"([OPQ][0-9][A-Z0-9]{3}[0-9]|[A-NR-Z][0-9]([A-Z][A-Z0-9]{2}[0-9]){1,2})"
# Ref.: https://www.uniprot.org/help/accession_numbers


def fetch_lipidmaps_info(query_id):
    """
    Recuperar información de cross-referencia e identificadores químicos de la API de LipidMaps.
    """
    if not query_id:
        return None
    # Determinar tipo de extensión para la consulta
    if query_id.startswith("LMP"):
        ext = "protein/lmp_id"
    elif re.match(UNIPROT_PATTERN, query_id):
        ext = "protein/uniprot_id"
    else:
        ext = "compound/lm_id"
    url = f"https://www.lipidmaps.org/rest/{ext}/{query_id}/all"

    try:
        r = requests.get(url, headers={"User-Agent": "Mozilla/5.0"}, timeout=10)
        r.raise_for_status()
        data = r.json()
        if isinstance(data, list) and not data:
            return None

        # Casos se manejan según formato de respuesta
        if isinstance(data, dict) and any(k.startswith('Row') for k in data):
            first_row = data.get('Row1') or {}
            return {
                "LM_ID": first_row.get("lm_id") or first_row.get("lmp_id") or query_id,
                "RefSeq_Id": first_row.get("refseq_id"),
                "UniProt": first_row.get("uniprot_id")
            }
        elif isinstance(data, dict):
            lm_from_data = data.get("lm_id") or data.get("lmp_id") or None
            return {
                "LM_ID": lm_from_data if lm_from_data else (query_id if query_id.startswith("LM") else None),
                "KEGG": data.get("kegg_id"),
                "PubChem": data.get("pubchem_cid"),
                "HMDB": data.get("hmdb_id"),
                "ChEBI": data.get("chebi_id"),
                "RefSeq_Id": data.get("refseq_id"),
                "UniProt": data.get("uniprot_id"),
                "InChIKey": data.get("inchi_key"),
                "InChI": data.get("inchi")
            }
        else:
            logging.error(f"Formato inesperado para {query_id}: {type(data)}")
            return None
    except (requests.exceptions.RequestException, ValueError) as e:
        logging.warning(f"No data for {query_id} ({e})")
        return None

def fetch_pubchem_info(pubchem_id):
    """
    Recuperar información de un compuesto a partir del ID PubChem (CID) usando la API REST.
    """
    try:
        cid = re.sub(r"[^0-9]", "", pubchem_id)
        if not cid:
            return None
        url_inchi = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/property/InChI,InChIKey/JSON"
        r_inchi = requests.get(url_inchi, timeout=10)
        r_inchi.raise_for_status()
        data_inchi = r_inchi.json().get("PropertyTable", {}).get("Properties", [{}])[0]
        # Recuperar cross-referencias
        url_xrefs = f"https://pubchem.ncbi.nlm.nih.gov/rest/pug/compound/cid/{cid}/xrefs/RegistryID/JSON"
        r_xrefs = requests.get(url_xrefs, timeout=10)
        r_xrefs.raise_for_status()
        data_xrefs = r_xrefs.json().get("InformationList", {}).get("Information", [{}])[0].get("RegistryID", [])
        kegg_id = next((x for x in data_xrefs if re.fullmatch(r"C\d{5}", x)), None)
        chebi_id = next((x for x in data_xrefs if x.startswith("CHEBI:")), None)
        hmdb_id = next((x for x in data_xrefs if x.startswith("HMDB")), None)
        return {
            "PubChem": cid,
            "KEGG": kegg_id,
            "HMDB": hmdb_id,
            "ChEBI": chebi_id,
            "InChI": data_inchi.get("InChI"),
            "InChIKey": data_inchi.get("InChIKey")
        }
    except Exception as e:
        logging.warning(f"PubChem no data for {pubchem_id}: {e}")
        return None

def fetch_refseq_from_uniprot(uniprot_id):
    """
    Recupera IDs de RefSeq cruzados desde UniProt (usando su API REST).
    """
    try:
        url = f"https://rest.uniprot.org/uniprotkb/{uniprot_id}?format=json"
        r = requests.get(url, timeout=10)
        r.raise_for_status()
        data = r.json()
        refseq_ids = []
        refseq_pairs = []
        for xref in data.get("uniProtKBCrossReferences", []):
            if xref.get("database") == "RefSeq":
                refseq_ids.append(xref.get("id"))
                nuc = None
                for prop in xref.get("properties", []):
                    if prop.get("key") == "NucleotideSequenceId":
                        nuc = prop.get("value")
                refseq_pairs.append((xref.get("id"), nuc))
        return {
            "UniProt": uniprot_id,
            "RefSeq_IDs": refseq_ids or None,
            "RefSeq_Pairs": refseq_pairs or None
        }
    except Exception as e:
        logging.warning(f"UniProt no data for {uniprot_id}: {e}")
        return None

def process_excel(input_file):
    """
    Procesa todas las hojas de un archivo Excel, consulta APIs externas y actualiza el Excel con las cross-referencias encontradas.
    """
    wb = load_workbook(input_file)
    total_time = []

    for sheet_name in wb.sheetnames:
        inicio = time.perf_counter()
        ws = wb[sheet_name]
        logging.info(f"--- Procesando hoja: {sheet_name} ---")
        df = pd.DataFrame(ws.values)
        df.columns = df.iloc[0]  # Definir encabezados
        df = df[2:]  # Saltar encabezado + nombre de pathway

        # Filtrado y conteo de IDs
        seen_queries = set()
        lm_ids, other_ids = [], []
        for id_raw in df["ID"]:
            id_str = str(id_raw).strip() if pd.notna(id_raw) else ''
            if not id_str:
                continue
            for sub in [s.strip() for s in id_str.split(";") if s.strip()]:
                if sub in seen_queries:
                    continue
                seen_queries.add(sub)
                if sub.startswith("LM"):
                    lm_ids.append(sub)
                else:
                    other_ids.append(sub)

        logging.info(f"Se encontraron {len(lm_ids)} IDs LM y {len(other_ids)} no-LM en {sheet_name}")

        # Consulta a LipidMaps y construcción del índice
        results = {}
        for query_id in (lm_ids + other_ids):
            info = fetch_lipidmaps_info(query_id)
            if not info:
                time.sleep(0.25)
                continue
            lm_key = info.get("LM_ID")
            if lm_key:
                results[lm_key] = info
            uni_field = info.get("UniProt")
            if uni_field:
                for uni in [u.strip() for u in str(uni_field).split(";") if u.strip()]:
                    results[uni] = info
            results[query_id] = info
            time.sleep(0.25)

        # IDs no crossreferenciadas
        not_crossreferenced = set()
        for id_raw in df["ID"]:
            id_str = str(id_raw).strip() if pd.notna(id_raw) else ''
            if not id_str:
                continue
            for sub in [s.strip() for s in id_str.split(";") if s.strip()]:
                if sub not in results:
                    not_crossreferenced.add(sub)

        # Consultar APIs externas (UniProt y PubChem)
        external_results = {}
        # Cambia y adapta lista de UniProt si procede
        uniprot_ids = ["Q06520", "Q14032", "Q9Y2P5"]
        for uid in uniprot_ids:
            info = fetch_refseq_from_uniprot(uid)
            if info:
                external_results[uid] = info
            time.sleep(0.3)
        for unknown_id in sorted(not_crossreferenced):
            if re.match(r"^\d+$", unknown_id):
                info = fetch_pubchem_info(unknown_id)
                if info:
                    external_results[unknown_id] = info
                time.sleep(0.3)
        results.update(external_results)

        # Excluir IDs UniProt del resumen de no cruzadas
        not_crossreferenced_final = [
            x for x in sorted(not_crossreferenced)
            if x not in results and not re.match(UNIPROT_PATTERN, x)
        ]
        if not_crossreferenced_final:
            logging.info(f"{len(not_crossreferenced_final)} IDs sin crossreference en {sheet_name}: {not_crossreferenced_final}")
        else:
            logging.info(f"Todas las IDs cruzadas en {sheet_name}")

        # Mapear columnas objetivo a índices
        header = [cell.value for cell in ws[1]]
        header_lower = [str(h).lower() if h else "" for h in header]
        target_col_idx = {}
        for col in TARGET_COLS:
            if col.lower() in header_lower:
                target_col_idx[col] = header_lower.index(col.lower()) + 1
        if "id" not in header_lower:
            logging.warning(f"No 'ID' header in hoja {sheet_name}. Saltando hoja.")
            continue
        id_idx = header_lower.index("id") + 1

        # Actualizar celdas con información encontrada
        for row in ws.iter_rows(min_row=2, values_only=False):
            id_cell = row[id_idx - 1]
            id_val = str(id_cell.value).strip() if pd.notna(id_cell.value) else ''
            if not id_val:
                continue
            sub_ids = [x.strip() for x in id_val.split(";") if x.strip()]
            merged_info = {k: [] for k in TARGET_COLS}
            new_id_val = []
            for sub_id in sub_ids:
                info = results.get(sub_id)
                if not info and re.match(UNIPROT_PATTERN, sub_id):
                    info = results.get(sub_id)
                if info:
                    lm_real = info.get("LM_ID")
                    new_id_val.append(lm_real if lm_real else sub_id)
                    for col in TARGET_COLS:
                        v = info.get(col)
                        if v:
                            for part in [p.strip() for p in str(v).split(";") if p.strip()]:
                                if part not in merged_info[col]:
                                    merged_info[col].append(part)
                else:
                    new_id_val.append(sub_id)
            # Escribir nuevos valores concatenados
            id_cell.value = ";".join(new_id_val)
            for col_name, col_idx in target_col_idx.items():
                vals = merged_info.get(col_name) or []
                if vals:
                    ws.cell(row=id_cell.row, column=col_idx, value=";".join(vals))
        fin = time.perf_counter()
        total_time.append(fin - inicio)
        logging.info(f"Tiempo de ejecución hoja: {fin - inicio:.2f} s")
    output_file = input_file.replace(".xlsx", "_updated.xlsx")
    wb.save(output_file)
    logging.info(f"Archivo guardado como: {output_file}")
    logging.info(f"Tiempo total: {sum(total_time):.2f} s")

# ------------------------------------------------------------------------------
if __name__ == "__main__":
    INPUT_FILE = "c:/Users/dborrotoa/Desktop/TFM/PathwayBA_list.xlsx"  # Adapta ruta si necesario
    process_excel(INPUT_FILE)
